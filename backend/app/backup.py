"""Back-up, export and restore.

Three artefacts come out of this module, and one goes back in:

* **Kastexport** — any user can download their own kast: an Excel workbook they
  can actually read, the photos as ordinary JPEG files, and a
  ``wardrobe.json`` that a restore reads back exactly.
* **Volledige export** (admin) — the same archive for every kast at once,
  including the accounts, the catalogue and the colour rules.
* **Momentopname** (admin) — the raw database file plus the uploads folder.
  Not portable between versions, but byte-exact: this is the one to keep for
  disaster recovery.

The Excel workbook is the *face* of an export and the JSON is its *truth*. A
spreadsheet is what someone opens on a rainy Sunday to look through their
wardrobe; the JSON is what the app reads when the archive has to become a
wardrobe again. Both are written from the same data in one pass, so they can
never disagree.

Garments are referenced by ``Item.uid`` rather than by row id. Ids are local to
one installation and get reused after a delete; a uid travels with the garment,
which is what makes a restore able to update instead of duplicate.
"""

from __future__ import annotations

import hashlib
import io
import json
import re
import shutil
import sqlite3
import tempfile
import unicodedata
import zipfile
from datetime import datetime, timezone
from pathlib import Path

from sqlalchemy.orm import Session

from ._version import __version__
from .config import settings
from .models import (
    Brand,
    Category,
    ColorRule,
    Item,
    Match,
    MatchSkip,
    SizeOption,
    User,
    Wardrobe,
    WardrobeMember,
    as_utc,
)

#: Marks an archive as ours. A restore refuses anything else outright.
FORMAT = "kledingkast-backup"
#: Bumped when the JSON layout changes in a way older builds cannot read.
FORMAT_VERSION = 1

WORKBOOK_NAME = "Kledingkast.xlsx"
DATA_NAME = "wardrobe.json"
MANIFEST_NAME = "manifest.json"
README_NAME = "LEESMIJ.txt"
PHOTO_DIR = "photos"

#: Photo shown inside the spreadsheet, in pixels. Small on purpose: a hundred
#: full-size photos would make the workbook heavier than the archive around it.
SHEET_PHOTO_PX = 96

VERDICT_LABELS = {"yes": "Past bij elkaar", "no": "Past niet", "skip": "Overgeslagen"}


# --------------------------------------------------------------------------
# gathering
# --------------------------------------------------------------------------


def _iso(value: datetime | None) -> str | None:
    return as_utc(value).isoformat() if value else None


def _person_key(user_id: int) -> str:
    return f"u{user_id}"


def _seasons(item: Item) -> list[str]:
    return [s.strip() for s in (item.season or "").split(",") if s.strip()]


def _photo_name(path: Path) -> str:
    """Name a photo after its content, so identical files are stored once.

    Duplicating a garment copies its photo on disk, so a kast can easily hold
    the same picture several times over. Hashing collapses those back into one
    archive entry — and gives a restore something to verify against.
    """
    digest = hashlib.sha256(path.read_bytes()).hexdigest()[:16]
    return f"{digest}.jpg"


def collect(
    db: Session,
    wardrobes: list[Wardrobe],
    *,
    scope: str,
    actor: User,
    with_accounts: bool = False,
) -> tuple[dict, dict[str, Path]]:
    """Build the export payload plus the photo files it refers to.

    Returns ``(payload, photos)`` where ``photos`` maps the name inside the
    archive to the file on disk.
    """
    wardrobe_ids = [w.id for w in wardrobes]
    items = (
        db.query(Item)
        .filter(Item.wardrobe_id.in_(wardrobe_ids))
        .order_by(Item.wardrobe_id, Item.name)
        .all()
        if wardrobe_ids
        else []
    )
    item_ids = [it.id for it in items]
    by_id = {it.id: it for it in items}

    def _pairs(model):
        if not item_ids:
            return []
        return (
            db.query(model)
            .filter(model.item_a_id.in_(item_ids), model.item_b_id.in_(item_ids))
            .all()
        )

    matches = _pairs(Match)
    skips = _pairs(MatchSkip)
    members = (
        db.query(WardrobeMember).filter(WardrobeMember.wardrobe_id.in_(wardrobe_ids)).all()
        if wardrobe_ids
        else []
    )

    # Everyone who appears anywhere in the file, so no reference dangles.
    person_ids = {w.owner_id for w in wardrobes}
    person_ids.update(m.user_id for m in members)
    person_ids.update(it.created_by_id for it in items)
    person_ids.update(m.user_id for m in matches)
    person_ids.update(s.user_id for s in skips)
    people = db.query(User).filter(User.id.in_(person_ids)).all() if person_ids else []

    photos: dict[str, Path] = {}

    def _photo_ref(item: Item) -> str | None:
        if not item.photo_filename:
            return None
        path = settings.uploads_dir / item.photo_filename
        if not path.exists():
            return None
        name = _photo_name(path)
        photos.setdefault(name, path)
        return f"{PHOTO_DIR}/{name}"

    items_by_wardrobe: dict[int, list[dict]] = {wid: [] for wid in wardrobe_ids}
    for item in items:
        items_by_wardrobe[item.wardrobe_id].append(
            {
                "uid": item.uid,
                "name": item.name,
                "category": item.category,
                "brand": item.brand,
                "color": item.color,
                "size": item.size,
                "seasons": _seasons(item),
                "notes": item.notes,
                "is_favorite": item.is_favorite,
                "created_at": _iso(item.created_at),
                "created_by": _person_key(item.created_by_id),
                "photo": _photo_ref(item),
            }
        )

    def _pair_rows(rows, verdict_of) -> dict[int, list[dict]]:
        out: dict[int, list[dict]] = {wid: [] for wid in wardrobe_ids}
        for row in rows:
            a, b = by_id[row.item_a_id], by_id[row.item_b_id]
            # A pair can straddle two kasten in a shared setup; file it under
            # the first garment's kast so it is exported exactly once.
            out[a.wardrobe_id].append(
                {
                    "items": [a.uid, b.uid],
                    "verdict": verdict_of(row),
                    "by": _person_key(row.user_id),
                    "decided_at": _iso(getattr(row, "updated_at", None) or row.created_at),
                }
            )
        return out

    combinations = _pair_rows(matches, lambda m: m.verdict)
    skipped = _pair_rows(skips, lambda _s: "skip")

    members_by_wardrobe: dict[int, list[dict]] = {wid: [] for wid in wardrobe_ids}
    for member in members:
        members_by_wardrobe[member.wardrobe_id].append(
            {"person": _person_key(member.user_id), "role": member.role}
        )

    payload = {
        "format": FORMAT,
        "format_version": FORMAT_VERSION,
        "app_version": __version__,
        "scope": scope,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "generated_by": actor.display_name,
        "people": [
            {
                "key": _person_key(p.id),
                "username": p.username,
                "display_name": p.display_name,
                **(
                    {"is_admin": p.is_admin, "hashed_password": p.hashed_password}
                    if with_accounts
                    else {}
                ),
            }
            for p in people
        ],
        "wardrobes": [
            {
                "key": f"w{w.id}",
                "name": w.name,
                "owner": _person_key(w.owner_id),
                "members": members_by_wardrobe.get(w.id, []),
                "items": items_by_wardrobe.get(w.id, []),
                "combinations": combinations.get(w.id, []),
                "skipped": skipped.get(w.id, []),
            }
            for w in wardrobes
        ],
    }

    if with_accounts:
        payload["catalog"] = {
            "categories": [
                {"name": c.name, "position": c.position}
                for c in db.query(Category).order_by(Category.position).all()
            ],
            "sizes": [
                {"label": s.label, "kind": s.kind, "position": s.position}
                for s in db.query(SizeOption).order_by(SizeOption.position).all()
            ],
            "color_rules": [
                {"color_a": r.color_a, "color_b": r.color_b, "verdict": r.verdict}
                for r in db.query(ColorRule).all()
            ],
        }

    return payload, photos


def counts_of(payload: dict) -> dict[str, int]:
    """The numbers a restore screen shows before touching anything."""
    wardrobes = payload.get("wardrobes", [])
    return {
        "wardrobes": len(wardrobes),
        "people": len(payload.get("people", [])),
        "items": sum(len(w.get("items", [])) for w in wardrobes),
        "combinations": sum(len(w.get("combinations", [])) for w in wardrobes),
        "skipped": sum(len(w.get("skipped", [])) for w in wardrobes),
    }


# --------------------------------------------------------------------------
# the spreadsheet
# --------------------------------------------------------------------------


def _thumb_bytes(path: Path) -> bytes | None:
    """A small JPEG of a photo, for showing inside the workbook."""
    from PIL import Image

    try:
        with Image.open(path) as img:
            img = img.convert("RGB")
            img.thumbnail((SHEET_PHOTO_PX, SHEET_PHOTO_PX))
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=75)
            return buf.getvalue()
    except Exception:  # a photo that will not open must not sink the export
        return None


def _autosize(ws, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def build_workbook(payload: dict, photos: dict[str, Path]) -> bytes:
    """Render the export as an Excel workbook people can read without us."""
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3B57")
    multi = len(payload.get("wardrobes", [])) > 1
    person_names = {p["key"]: p["display_name"] for p in payload.get("people", [])}

    # ---- Lees mij -------------------------------------------------------
    intro = wb.active
    intro.title = "Lees mij"
    counts = counts_of(payload)
    lines = [
        ("Kledingkast — export", True),
        ("", False),
        (f"Gemaakt op: {payload['generated_at'][:19].replace('T', ' ')} (UTC)", False),
        (f"Gemaakt door: {payload['generated_by']}", False),
        (f"App-versie: {payload['app_version']}", False),
        ("", False),
        (f"Kledingstukken: {counts['items']}", False),
        (f"Combinaties: {counts['combinations']}", False),
        (f"Kasten: {counts['wardrobes']}", False),
        ("", False),
        ("Wat zit hier in?", True),
        ("• Tabblad 'Kledingstukken' — alles wat je hebt vastgelegd, met een foto per regel.", False),
        ("• Tabblad 'Combinaties' — welke stukken bij elkaar passen, en van wie dat oordeel is.", False),
        ("• Map 'photos' — de foto's zelf, als gewone JPEG-bestanden.", False),
        (f"• Bestand '{DATA_NAME}' — dezelfde gegevens in machineleesbare vorm.", False),
        ("", False),
        ("Tip: pak het hele ZIP-bestand uit voordat je dit opent.", True),
        ("De links naar de foto's werken alleen als de map 'photos' naast dit bestand staat.", False),
        ("", False),
        ("Sorteer of filter je de lijst? De foto's blijven dan staan waar ze staan —", False),
        ("dat doet Excel nu eenmaal met afbeeldingen. De kolom 'Fotobestand' klopt altijd.", False),
        ("", False),
        (f"Terugzetten in de app doet een beheerder met het hele ZIP-bestand; daarvoor is '{DATA_NAME}' nodig.", False),
    ]
    for row, (text, bold) in enumerate(lines, start=1):
        cell = intro.cell(row=row, column=1, value=text)
        if bold:
            cell.font = Font(bold=True)
    _autosize(intro, {"A": 100})

    # ---- Kledingstukken -------------------------------------------------
    sheet = wb.create_sheet("Kledingstukken")
    headers = ["Foto", "Naam", "Categorie", "Merk", "Kleur", "Maat", "Seizoen",
               "Favoriet", "Notities", "Toegevoegd op", "Toegevoegd door", "Fotobestand"]
    if multi:
        headers.insert(1, "Kast")
    headers.append("Kenmerk (niet wijzigen)")
    for col, title in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
    sheet.freeze_panes = "A2"

    photo_col = 1
    file_col = headers.index("Fotobestand") + 1
    row = 2
    for wardrobe in payload.get("wardrobes", []):
        for item in wardrobe.get("items", []):
            values = [
                None,  # the photo is placed, not written
                item["name"],
                item["category"],
                item["brand"],
                item["color"],
                item["size"],
                ", ".join(item["seasons"]),
                "ja" if item["is_favorite"] else "",
                item["notes"],
                (item["created_at"] or "")[:10],
                person_names.get(item["created_by"], ""),
                item["photo"] or "",
                item["uid"],
            ]
            if multi:
                values.insert(1, wardrobe["name"])
            for col, value in enumerate(values, start=1):
                cell = sheet.cell(row=row, column=col, value=value)
                cell.alignment = Alignment(vertical="center", wrap_text=(col == len(values) - 4))

            if item["photo"]:
                name = item["photo"].split("/", 1)[1]
                source = photos.get(name)
                raw = _thumb_bytes(source) if source else None
                if raw:
                    image = XLImage(io.BytesIO(raw))
                    image.anchor = f"{get_column_letter(photo_col)}{row}"
                    sheet.add_image(image)
                link = sheet.cell(row=row, column=file_col)
                # Relative link: works once the archive is unpacked as a whole.
                link.hyperlink = item["photo"]
                link.style = "Hyperlink"
            sheet.row_dimensions[row].height = SHEET_PHOTO_PX * 0.78
            row += 1

    widths = {"A": 15, "B": 28, "C": 14, "D": 16, "E": 14, "F": 10, "G": 20,
              "H": 9, "I": 34, "J": 14, "K": 16, "L": 26, "M": 34}
    if multi:
        widths = {"A": 15, "B": 20, "C": 28, "D": 14, "E": 16, "F": 14, "G": 10,
                  "H": 20, "I": 9, "J": 34, "K": 14, "L": 16, "M": 26, "N": 34}
    _autosize(sheet, widths)

    # ---- Combinaties ----------------------------------------------------
    pairs = wb.create_sheet("Combinaties")
    pair_headers = ["Stuk", "Combineert met", "Oordeel", "Volgens", "Wanneer"]
    if multi:
        pair_headers.insert(0, "Kast")
    pair_headers += ["Kenmerk A", "Kenmerk B"]
    for col, title in enumerate(pair_headers, start=1):
        cell = pairs.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
    pairs.freeze_panes = "A2"

    names = {
        item["uid"]: item["name"]
        for wardrobe in payload.get("wardrobes", [])
        for item in wardrobe.get("items", [])
    }
    prow = 2
    for wardrobe in payload.get("wardrobes", []):
        rows = list(wardrobe.get("combinations", [])) + list(wardrobe.get("skipped", []))
        for entry in sorted(rows, key=lambda e: names.get(e["items"][0], "")):
            uid_a, uid_b = entry["items"]
            values = [
                names.get(uid_a, uid_a),
                names.get(uid_b, uid_b),
                VERDICT_LABELS.get(entry["verdict"], entry["verdict"]),
                person_names.get(entry["by"], ""),
                (entry["decided_at"] or "")[:10],
                uid_a,
                uid_b,
            ]
            if multi:
                values.insert(0, wardrobe["name"])
            for col, value in enumerate(values, start=1):
                pairs.cell(row=prow, column=col, value=value)
            prow += 1
    _autosize(pairs, dict(zip("ABCDEFGH", ([20] if multi else []) + [28, 28, 18, 16, 14, 34, 34])))

    # ---- Gebruikers (alleen in een volledige export) ---------------------
    if payload.get("scope") == "instance":
        users = wb.create_sheet("Gebruikers")
        for col, title in enumerate(["Naam", "Gebruikersnaam", "Beheerder"], start=1):
            cell = users.cell(row=1, column=col, value=title)
            cell.font = header_font
            cell.fill = header_fill
        for urow, person in enumerate(payload.get("people", []), start=2):
            users.cell(row=urow, column=1, value=person["display_name"])
            users.cell(row=urow, column=2, value=person["username"])
            users.cell(row=urow, column=3, value="ja" if person.get("is_admin") else "")
        _autosize(users, {"A": 24, "B": 20, "C": 12})

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# --------------------------------------------------------------------------
# writing the archive
# --------------------------------------------------------------------------


def safe_filename(name: str, suffix: str) -> str:
    """A download name that survives every browser and filesystem."""
    plain = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode()
    slug = re.sub(r"[^A-Za-z0-9]+", "-", plain).strip("-").lower() or "kledingkast"
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    return f"kledingkast-{slug}-{stamp}{suffix}"


def _readme(payload: dict) -> str:
    counts = counts_of(payload)
    return "\n".join(
        [
            "Kledingkast — export",
            "",
            f"Gemaakt op   : {payload['generated_at'][:19].replace('T', ' ')} UTC",
            f"Gemaakt door : {payload['generated_by']}",
            f"App-versie   : {payload['app_version']}",
            f"Inhoud       : {counts['items']} kledingstukken, "
            f"{counts['combinations']} combinaties, {counts['wardrobes']} kast(en)",
            "",
            f"{WORKBOOK_NAME}",
            "    Open dit in Excel, LibreOffice of Numbers. Elke regel is een kledingstuk,",
            "    met een kleine foto erbij en een link naar het volledige bestand. Sorteer je",
            "    de lijst, dan blijven de foto's staan waar ze staan; de kolom 'Fotobestand'",
            "    blijft wel kloppen.",
            "",
            f"{PHOTO_DIR}/",
            "    De foto's zelf, als gewone JPEG-bestanden.",
            "",
            f"{DATA_NAME}",
            "    Dezelfde gegevens, machineleesbaar. Dit bestand heeft een beheerder nodig",
            "    om de kast in de app terug te zetten. Niet handmatig aanpassen.",
            "",
            "Pak het ZIP-bestand in zijn geheel uit: de links in het Excel-bestand",
            "wijzen naar de map photos ernaast.",
            "",
        ]
    )


def write_archive(payload: dict, photos: dict[str, Path]) -> Path:
    """Write the export to a temporary ZIP and return its path.

    Written to disk rather than held in memory: a kast of a few hundred photos
    is tens of megabytes, and the response streams straight from this file.
    """
    handle = tempfile.NamedTemporaryFile(prefix="kledingkast-export-", suffix=".zip", delete=False)
    handle.close()
    path = Path(handle.name)
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(
            MANIFEST_NAME,
            json.dumps(
                {
                    "format": FORMAT,
                    "format_version": FORMAT_VERSION,
                    "app_version": payload["app_version"],
                    "scope": payload["scope"],
                    "generated_at": payload["generated_at"],
                    "generated_by": payload["generated_by"],
                    "counts": counts_of(payload),
                },
                ensure_ascii=False,
                indent=2,
            ),
        )
        archive.writestr(README_NAME, _readme(payload))
        archive.writestr(DATA_NAME, json.dumps(payload, ensure_ascii=False, indent=1))
        archive.writestr(WORKBOOK_NAME, build_workbook(payload, photos))
        for name, source in photos.items():
            # Already-compressed JPEG: storing it saves time and gains nothing
            # by being deflated again.
            archive.write(source, f"{PHOTO_DIR}/{name}", compress_type=zipfile.ZIP_STORED)
    return path


def write_snapshot() -> Path:
    """The raw database plus the uploads folder, for disaster recovery.

    The database is copied through SQLite's own backup API rather than as a
    file copy: the app is serving requests while this runs, and a plain copy of
    a live database can catch it mid-write.
    """
    handle = tempfile.NamedTemporaryFile(prefix="kledingkast-snapshot-", suffix=".zip", delete=False)
    handle.close()
    path = Path(handle.name)

    with tempfile.TemporaryDirectory(prefix="kledingkast-db-") as tmp:
        db_copy = Path(tmp) / "wardrobe.db"
        source = sqlite3.connect(str(settings.db_path))
        try:
            target = sqlite3.connect(str(db_copy))
            try:
                source.backup(target)
            finally:
                target.close()
        finally:
            source.close()

        with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
            archive.writestr(
                MANIFEST_NAME,
                json.dumps(
                    {
                        "format": FORMAT,
                        "format_version": FORMAT_VERSION,
                        "app_version": __version__,
                        "scope": "snapshot",
                        "generated_at": datetime.now(timezone.utc).isoformat(),
                    },
                    ensure_ascii=False,
                    indent=2,
                ),
            )
            archive.writestr(README_NAME, SNAPSHOT_README)
            archive.write(db_copy, "wardrobe.db")
            for photo in sorted(settings.uploads_dir.glob("*")):
                if photo.is_file():
                    archive.write(photo, f"uploads/{photo.name}", compress_type=zipfile.ZIP_STORED)
    return path


SNAPSHOT_README = """Kledingkast — momentopname
==========================

Dit is een exacte kopie van de database en alle foto's, gemaakt terwijl de app
draaide. Bedoeld om na een crash of een verkeerde ingreep alles terug te zetten
zoals het was.

Terugzetten (docker compose):

    docker compose down
    # pak dit archief uit en zet de bestanden in het datavolume:
    #   wardrobe.db  ->  /data/wardrobe.db
    #   uploads/*    ->  /data/uploads/
    docker compose up -d

Let op: dit hoort bij dezelfde app-versie als waarmee het gemaakt is. Voor een
overdraagbare back-up (die ook een versie later nog te lezen is) gebruik je de
volledige export uit Instellingen.
"""


# --------------------------------------------------------------------------
# reading an archive back in
# --------------------------------------------------------------------------

#: Caps on what a restore will unpack. An archive is an untrusted file, even
#: when a beheerder uploaded it: these keep a crafted ZIP from filling the disk.
MAX_ENTRIES = 20_000
MAX_UNPACKED_BYTES = 2 * 1024 * 1024 * 1024  # 2 GB
MAX_JSON_BYTES = 64 * 1024 * 1024


class RestoreError(Exception):
    """The archive cannot be read. The message is shown to the user as-is."""


def _member_path(name: str) -> str:
    """Reject anything that would write outside the archive's own tree."""
    if name.startswith("/") or name.startswith("\\") or ".." in Path(name).parts:
        raise RestoreError(f"Onveilig pad in het archief: {name}")
    return name


def read_archive(path: Path) -> tuple[dict, dict[str, bytes]]:
    """Validate an uploaded archive and return ``(payload, photos)``.

    Nothing is written to the database here; this is what the confirmation
    screen calls to say what the file contains before anything happens.
    """
    try:
        archive = zipfile.ZipFile(path)
    except zipfile.BadZipFile:
        raise RestoreError("Dit is geen geldig ZIP-bestand.")

    with archive:
        entries = archive.infolist()
        if len(entries) > MAX_ENTRIES:
            raise RestoreError("Het archief bevat onwerkbaar veel bestanden.")
        total = 0
        for entry in entries:
            _member_path(entry.filename)
            total += entry.file_size
            if total > MAX_UNPACKED_BYTES:
                raise RestoreError("Het archief is uitgepakt te groot.")

        names = {entry.filename for entry in entries}
        if DATA_NAME not in names:
            if "wardrobe.db" in names:
                raise RestoreError(
                    "Dit is een momentopname. Die zet je terug op de server, "
                    "niet via de app — zie LEESMIJ.txt in het archief."
                )
            raise RestoreError(f"Het archief mist {DATA_NAME}; dit is geen kledingkast-export.")

        info = archive.getinfo(DATA_NAME)
        if info.file_size > MAX_JSON_BYTES:
            raise RestoreError("Het gegevensbestand in het archief is onwerkbaar groot.")
        try:
            payload = json.loads(archive.read(DATA_NAME))
        except (json.JSONDecodeError, UnicodeDecodeError):
            raise RestoreError(f"{DATA_NAME} is beschadigd of onleesbaar.")

        if not isinstance(payload, dict) or payload.get("format") != FORMAT:
            raise RestoreError("Dit archief komt niet uit Kledingkast.")
        version = payload.get("format_version")
        if not isinstance(version, int) or version > FORMAT_VERSION:
            raise RestoreError(
                f"Dit archief is gemaakt met een nieuwere versie (formaat {version}). "
                "Werk de app eerst bij."
            )
        if not isinstance(payload.get("wardrobes"), list):
            raise RestoreError(f"{DATA_NAME} mist de kastgegevens.")

        photos: dict[str, bytes] = {}
        for entry in entries:
            if entry.filename.startswith(f"{PHOTO_DIR}/") and not entry.is_dir():
                photos[entry.filename.split("/", 1)[1]] = archive.read(entry)

    return payload, photos


def _person_lookup(db: Session, payload: dict, *, fallback: User) -> dict[str, User]:
    """Map the file's people onto real accounts.

    Matched on username, which is unique. Someone in the file without an
    account here (a partner who judged garments on an installation you are
    restoring from) collapses onto ``fallback`` — their verdicts are kept, the
    account is not silently created.
    """
    mapping: dict[str, User] = {}
    for person in payload.get("people", []):
        username = (person.get("username") or "").strip().lower()
        user = db.query(User).filter(User.username == username).first() if username else None
        mapping[person.get("key", "")] = user or fallback
    return mapping


def apply_payload(
    db: Session,
    payload: dict,
    photos: dict[str, bytes],
    *,
    mode: str,
    target: Wardrobe,
    actor: User,
) -> dict[str, int]:
    """Write an archive's contents into ``target``. Caller commits.

    ``mode`` is ``"merge"`` (add and update, never delete) or ``"replace"``
    (empty the kast first). Garments are matched on uid, so restoring the same
    file twice leaves the same wardrobe, not two of everything.
    """
    from .accounts import clear_wardrobe_items
    from .images import store_image_bytes

    if mode not in {"merge", "replace"}:
        raise RestoreError("Onbekende herstelmodus.")

    people = _person_lookup(db, payload, fallback=actor)
    stats = {"added": 0, "updated": 0, "combinations": 0, "skipped_pairs": 0, "photos": 0}

    if mode == "replace":
        clear_wardrobe_items(db, target)
        db.flush()

    existing = {
        item.uid: item
        for item in db.query(Item).filter(Item.wardrobe_id == target.id).all()
    }
    by_uid: dict[str, Item] = dict(existing)

    for wardrobe in payload.get("wardrobes", []):
        for row in wardrobe.get("items", []):
            uid = row.get("uid")
            if not uid or not row.get("name"):
                continue
            item = by_uid.get(uid)
            if item is None:
                item = Item(uid=uid, wardrobe_id=target.id, created_by_id=actor.id)
                db.add(item)
                stats["added"] += 1
            else:
                stats["updated"] += 1
            item.name = row["name"]
            item.category = row.get("category") or "Overig"
            item.brand_ref = _brand(db, row.get("brand"))
            item.color = row.get("color")
            item.size = row.get("size")
            item.season = ", ".join(row.get("seasons") or []) or None
            item.notes = row.get("notes")
            item.is_favorite = bool(row.get("is_favorite"))
            creator = people.get(row.get("created_by", ""))
            item.created_by_id = (creator or actor).id

            photo_ref = row.get("photo")
            if photo_ref and not item.photo_filename:
                raw = photos.get(photo_ref.split("/", 1)[-1])
                if raw:
                    try:
                        item.photo_filename, item.thumb_filename = store_image_bytes(raw)
                        stats["photos"] += 1
                    except Exception:
                        # A corrupt photo costs its picture, never the garment.
                        item.photo_filename = item.thumb_filename = None
            db.flush()
            by_uid[uid] = item

    # Verdicts come last: both garments of a pair must exist by now.
    for wardrobe in payload.get("wardrobes", []):
        for entry in wardrobe.get("combinations", []) + wardrobe.get("skipped", []):
            pair = entry.get("items") or []
            if len(pair) != 2:
                continue
            first, second = by_uid.get(pair[0]), by_uid.get(pair[1])
            if not first or not second or first.id == second.id:
                continue
            a_id, b_id = sorted((first.id, second.id))
            judge = people.get(entry.get("by", "")) or actor
            verdict = entry.get("verdict")
            if verdict == "skip":
                if not _has(db, MatchSkip, a_id, b_id, judge.id):
                    db.add(MatchSkip(item_a_id=a_id, item_b_id=b_id, user_id=judge.id))
                    stats["skipped_pairs"] += 1
            elif verdict in {"yes", "no"}:
                row = _has(db, Match, a_id, b_id, judge.id)
                if row is None:
                    db.add(Match(item_a_id=a_id, item_b_id=b_id, user_id=judge.id, verdict=verdict))
                    stats["combinations"] += 1
                else:
                    row.verdict = verdict
    db.flush()
    return stats


def _has(db: Session, model, a_id: int, b_id: int, user_id: int):
    return (
        db.query(model)
        .filter(model.item_a_id == a_id, model.item_b_id == b_id, model.user_id == user_id)
        .first()
    )


def _brand(db: Session, name: str | None) -> Brand | None:
    """Resolve a brand name to the shared row, creating it when new."""
    clean = (name or "").strip()
    if not clean:
        return None
    brand = db.query(Brand).filter(Brand.name == clean).first()
    if brand is None:
        brand = Brand(name=clean)
        db.add(brand)
        db.flush()
    return brand
