"""Exporting a kast, and putting one back.

The export is two things at once: a spreadsheet a person reads, and a JSON
document the app reads. These tests hold both to their promises — the workbook
lists every garment with a link to its photo, the JSON survives a round trip
through a restore, and an archive from anywhere else is refused.
"""

import io
import json
import uuid
import zipfile

import pytest
from openpyxl import load_workbook
from PIL import Image

ADMIN_USER = "admin"
ADMIN_PASS = "changeme"


def h(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


def login(client, username: str, password: str) -> str:
    r = client.post("/api/auth/login", data={"username": username, "password": password})
    assert r.status_code == 200, r.text
    return r.json()["access_token"]


def make_user(client, admin_token: str, display: str):
    username = "u_" + uuid.uuid4().hex[:10]
    password = "pw123456"
    r = client.post(
        "/api/users",
        headers=h(admin_token),
        json={"username": username, "display_name": display, "password": password, "is_admin": False},
    )
    assert r.status_code == 201, r.text
    return username, password


def own_wardrobe_id(client, token: str) -> int:
    r = client.get("/api/wardrobes", headers=h(token))
    return [w for w in r.json() if w["my_role"] == "owner"][0]["id"]


def _jpeg(color=(120, 140, 160)) -> bytes:
    buf = io.BytesIO()
    Image.new("RGB", (60, 80), color).save(buf, "JPEG")
    return buf.getvalue()


def add_item(client, token, wardrobe_id, name, category, *, photo=True, **extra):
    data = {"wardrobe_id": str(wardrobe_id), "name": name, "category": category}
    data.update({k: str(v) for k, v in extra.items()})
    files = {"photo": ("p.jpg", _jpeg(), "image/jpeg")} if photo else None
    r = client.post("/api/items", headers=h(token), data=data, files=files)
    assert r.status_code == 201, r.text
    return r.json()


def export_zip(client, token, wardrobe_id) -> zipfile.ZipFile:
    r = client.get(f"/api/backup/wardrobe/{wardrobe_id}", headers=h(token))
    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == "application/zip"
    return zipfile.ZipFile(io.BytesIO(r.content))


@pytest.fixture
def stocked(client):
    """A kast with two garments, a photo each, and a verdict on the pair."""
    admin = login(client, ADMIN_USER, ADMIN_PASS)
    un, pw = make_user(client, admin, "Alice")
    token = login(client, un, pw)
    wid = own_wardrobe_id(client, token)
    polo = add_item(client, token, wid, "Witte polo", "Polo", brand="State of Art", color="Wit")
    jeans = add_item(client, token, wid, "Blauwe jeans", "Jeans", color="Blauw", photo=False)
    r = client.post(
        "/api/matches",
        headers=h(token),
        json={"item_a_id": polo["id"], "item_b_id": jeans["id"], "verdict": "yes"},
    )
    assert r.status_code == 204, r.text
    return {"admin": admin, "token": token, "wid": wid, "polo": polo, "jeans": jeans}


def test_export_carries_workbook_photos_and_data(client, stocked):
    with export_zip(client, stocked["token"], stocked["wid"]) as archive:
        names = archive.namelist()
        assert "Kledingkast.xlsx" in names
        assert "wardrobe.json" in names
        assert "manifest.json" in names
        assert "LEESMIJ.txt" in names
        # One garment has a photo, the other does not.
        assert len([n for n in names if n.startswith("photos/")]) == 1

        manifest = json.loads(archive.read("manifest.json"))
        assert manifest["format"] == "kledingkast-backup"
        assert manifest["counts"]["items"] == 2
        assert manifest["counts"]["combinations"] == 1

        payload = json.loads(archive.read("wardrobe.json"))
        items = payload["wardrobes"][0]["items"]
        assert {it["name"] for it in items} == {"Witte polo", "Blauwe jeans"}
        polo = next(it for it in items if it["name"] == "Witte polo")
        assert polo["brand"] == "State of Art"
        assert polo["photo"].startswith("photos/")
        # The pair is referenced by uid, never by database id.
        combo = payload["wardrobes"][0]["combinations"][0]
        assert set(combo["items"]) == {it["uid"] for it in items}
        assert combo["verdict"] == "yes"

        book = load_workbook(io.BytesIO(archive.read("Kledingkast.xlsx")))
        assert book.sheetnames[:3] == ["Lees mij", "Kledingstukken", "Combinaties"]
        sheet = book["Kledingstukken"]
        assert sheet.cell(row=1, column=2).value == "Naam"
        rows = {sheet.cell(row=r, column=2).value for r in range(2, sheet.max_row + 1)}
        assert rows == {"Witte polo", "Blauwe jeans"}
        # The photo is both linked and shown.
        link_col = [c.value for c in sheet[1]].index("Fotobestand") + 1
        links = [
            sheet.cell(row=r, column=link_col).hyperlink
            for r in range(2, sheet.max_row + 1)
        ]
        assert any(link is not None and link.target.startswith("photos/") for link in links)
        assert len(sheet._images) == 1

        pairs = book["Combinaties"]
        assert pairs.cell(row=2, column=3).value == "Past bij elkaar"


def test_export_includes_other_peoples_verdicts(client, stocked):
    """A shared kast collects opinions; an export keeps whose they were."""
    admin, token, wid = stocked["admin"], stocked["token"], stocked["wid"]
    un, pw = make_user(client, admin, "Anne")
    guest = login(client, un, pw)
    assert client.post(
        f"/api/wardrobes/{wid}/members",
        headers=h(token),
        json={"username": un, "role": "viewer"},
    ).status_code == 201

    assert client.post(
        "/api/matches",
        headers=h(guest),
        json={
            "item_a_id": stocked["polo"]["id"],
            "item_b_id": stocked["jeans"]["id"],
            "verdict": "no",
        },
    ).status_code == 204

    with export_zip(client, token, wid) as archive:
        payload = json.loads(archive.read("wardrobe.json"))
    verdicts = {c["verdict"] for c in payload["wardrobes"][0]["combinations"]}
    assert verdicts == {"yes", "no"}
    names = {p["display_name"] for p in payload["people"]}
    assert {"Alice", "Anne"} <= names


def test_restore_is_admin_only_and_round_trips(client, stocked):
    admin, token, wid = stocked["admin"], stocked["token"], stocked["wid"]
    r = client.get(f"/api/backup/wardrobe/{wid}", headers=h(token))
    blob = r.content

    # An ordinary user cannot restore, inspect or take a full backup.
    for path in ("/api/backup/instance", "/api/backup/snapshot"):
        assert client.get(path, headers=h(token)).status_code == 403
    assert client.post(
        "/api/backup/inspect", headers=h(token), files={"file": ("b.zip", blob, "application/zip")}
    ).status_code == 403

    # The beheerder first asks what is in the file.
    r = client.post(
        "/api/backup/inspect", headers=h(admin), files={"file": ("b.zip", blob, "application/zip")}
    )
    assert r.status_code == 200, r.text
    assert r.json()["items"] == 2
    assert r.json()["scope"] == "wardrobe"

    # Restoring into a *different* kast recreates both garments and the verdict.
    other = own_wardrobe_id(client, admin)
    r = client.post(
        "/api/backup/restore",
        headers=h(admin),
        files={"file": ("b.zip", blob, "application/zip")},
        data={"wardrobe_id": str(other), "mode": "merge"},
    )
    assert r.status_code == 200, r.text
    assert r.json()["added"] == 2
    assert r.json()["combinations"] == 1

    items = client.get(f"/api/items?wardrobe_id={other}", headers=h(admin)).json()
    assert {it["name"] for it in items} == {"Witte polo", "Blauwe jeans"}
    assert any(it["photo_filename"] for it in items)

    # Restoring the very same file again updates instead of duplicating.
    r = client.post(
        "/api/backup/restore",
        headers=h(admin),
        files={"file": ("b.zip", blob, "application/zip")},
        data={"wardrobe_id": str(other), "mode": "merge"},
    )
    assert r.status_code == 200, r.text
    assert r.json() == {**r.json(), "added": 0, "updated": 2}
    again = client.get(f"/api/items?wardrobe_id={other}", headers=h(admin)).json()
    assert len(again) == len(items)


def test_replace_empties_the_kast_first(client, stocked):
    admin, token, wid = stocked["admin"], stocked["token"], stocked["wid"]
    blob = client.get(f"/api/backup/wardrobe/{wid}", headers=h(token)).content

    target = own_wardrobe_id(client, admin)
    stray = add_item(client, admin, target, "Oude trui", "Trui", photo=False)

    r = client.post(
        "/api/backup/restore",
        headers=h(admin),
        files={"file": ("b.zip", blob, "application/zip")},
        data={"wardrobe_id": str(target), "mode": "replace"},
    )
    assert r.status_code == 200, r.text
    names = {it["name"] for it in client.get(f"/api/items?wardrobe_id={target}", headers=h(admin)).json()}
    assert "Oude trui" not in names
    assert names == {"Witte polo", "Blauwe jeans"}
    # The garment is really gone, not merely hidden. (SQLite hands a freed row
    # id straight to the next insert, so the id may well exist again — under a
    # different name.)
    stale = client.get(f"/api/items/{stray['id']}", headers=h(admin))
    assert stale.status_code == 404 or stale.json()["name"] != "Oude trui"


def test_a_foreign_or_hostile_archive_is_refused(client):
    admin = login(client, ADMIN_USER, ADMIN_PASS)
    wid = own_wardrobe_id(client, admin)

    def post(blob, path="/api/backup/inspect", **extra):
        return client.post(
            path, headers=h(admin), files={"file": ("x.zip", blob, "application/zip")}, **extra
        )

    assert post(b"not a zip at all").status_code == 400

    # A ZIP without our data file.
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("random.txt", "hallo")
    assert post(buf.getvalue()).status_code == 400

    # A path that would escape the archive when unpacked.
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("wardrobe.json", json.dumps({"format": "kledingkast-backup",
                                                "format_version": 1, "wardrobes": []}))
        z.writestr("../../etc/passwd", "root")
    r = post(buf.getvalue())
    assert r.status_code == 400
    assert "Onveilig pad" in r.json()["detail"]

    # A file from a newer format version we cannot promise to understand.
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("wardrobe.json", json.dumps({"format": "kledingkast-backup",
                                                "format_version": 99, "wardrobes": []}))
    assert post(buf.getvalue()).status_code == 400

    # And a snapshot, which is restored on the server rather than through here.
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("wardrobe.db", b"SQLite format 3\x00")
    r = post(buf.getvalue())
    assert r.status_code == 400
    assert "momentopname" in r.json()["detail"].lower()


def test_full_backup_and_snapshot_are_downloadable_by_admin(client, stocked):
    admin = stocked["admin"]

    r = client.get("/api/backup/instance", headers=h(admin))
    assert r.status_code == 200, r.text
    with zipfile.ZipFile(io.BytesIO(r.content)) as archive:
        payload = json.loads(archive.read("wardrobe.json"))
        assert payload["scope"] == "instance"
        assert len(payload["wardrobes"]) >= 2
        assert "catalog" in payload
        book = load_workbook(io.BytesIO(archive.read("Kledingkast.xlsx")))
        assert "Gebruikers" in book.sheetnames
        # A full backup carries accounts, so the kast column tells them apart.
        assert book["Kledingstukken"].cell(row=1, column=2).value == "Kast"

    r = client.get("/api/backup/snapshot", headers=h(admin))
    assert r.status_code == 200, r.text
    with zipfile.ZipFile(io.BytesIO(r.content)) as archive:
        assert "wardrobe.db" in archive.namelist()
        assert archive.read("wardrobe.db")[:15] == b"SQLite format 3"
        assert any(n.startswith("uploads/") for n in archive.namelist())
